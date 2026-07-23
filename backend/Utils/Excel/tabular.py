"""读写 .xls / .xlsx，自动探测表头行。"""

from __future__ import annotations

from pathlib import Path
from typing import Any


def _norm(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def read_tabular(path: str | Path, *, sheet_index: int = 0, sheet_name: str | None = None) -> list[dict[str, str]]:
    """返回 [{header: value}, ...]。跳过空行。"""

    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(path, sheet_index=sheet_index, sheet_name=sheet_name)
    if suffix == ".xls":
        return _read_xls(path, sheet_index=sheet_index, sheet_name=sheet_name)
    raise ValueError(f"unsupported file type: {path}")


def _detect_header_row(rows: list[list[str]], keywords: tuple[str, ...] = ("学号", "姓名", "年级")) -> int:
    best_i, best_score = 0, -1
    for i, row in enumerate(rows[:15]):
        joined = "".join(row)
        score = sum(1 for k in keywords if k in joined)
        nonempty = sum(1 for c in row if c)
        score = score * 10 + min(nonempty, 20)
        if score > best_score:
            best_i, best_score = i, score
    return best_i


def _rows_to_dicts(matrix: list[list[str]]) -> list[dict[str, str]]:
    if not matrix:
        return []
    hi = _detect_header_row(matrix)
    headers = [h or f"col_{i}" for i, h in enumerate(matrix[hi])]
    # 去重表头
    seen: dict[str, int] = {}
    uniq_headers: list[str] = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            uniq_headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            uniq_headers.append(h)

    out: list[dict[str, str]] = []
    for row in matrix[hi + 1 :]:
        if not any(row):
            continue
        item = {uniq_headers[i]: (row[i] if i < len(row) else "") for i in range(len(uniq_headers))}
        if not any(item.values()):
            continue
        out.append(item)
    return out


def _read_xlsx(path: Path, *, sheet_index: int, sheet_name: str | None) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.worksheets[sheet_index]
        matrix: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            matrix.append([_norm(c) for c in row])
        return _rows_to_dicts(matrix)
    finally:
        wb.close()


def _read_xls(path: Path, *, sheet_index: int, sheet_name: str | None) -> list[dict[str, str]]:
    import xlrd

    book = xlrd.open_workbook(str(path))
    sh = book.sheet_by_name(sheet_name) if sheet_name else book.sheet_by_index(sheet_index)
    matrix: list[list[str]] = []
    for r in range(sh.nrows):
        matrix.append([_norm(sh.cell_value(r, c)) for c in range(sh.ncols)])
    return _rows_to_dicts(matrix)


def list_sheet_names(path: str | Path) -> list[str]:
    path = Path(path)
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    import xlrd

    return list(xlrd.open_workbook(str(path)).sheet_names())
