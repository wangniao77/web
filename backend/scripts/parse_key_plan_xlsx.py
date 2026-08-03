"""解析《学院重点工作进展.xlsx》→ 结构化分组任务。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

SECTION_MAP: dict[str, tuple[str, str, str]] = {
    "学科建设": ("discipline", "学科建设", "学院发展根基"),
    "师资建设": ("faculty", "师资队伍建设", "学院发展命脉"),
    "教学建设": ("teaching", "教学建设", "人才培养主阵地"),
    "科研建设": ("research", "科研建设", "创新驱动引擎"),
    "人才培养": ("talent", "人才培养", "立德树人核心"),
    "党建办学": ("party", "党建与综合办学保障", "政治引领与办学保障"),
    "广财AI智教": ("ai", "广财AI智教专项改革", "数字化转型专项"),
}


def _raw(ws, row: int, col: int) -> Any:
    return ws.cell(row, col).value


def _merged(ws, row: int, col: int) -> Any:
    v = ws.cell(row, col).value
    if v is not None:
        return v
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return ws.cell(mr.min_row, mr.min_col).value
    return None


def _parse_section(text: str) -> tuple[str, str, str, str] | None:
    text = str(text).strip()
    owner = ""
    title = text
    for sep in ("——", "—", "-"):
        if sep in text:
            title, owner = text.split(sep, 1)
            break
    title = title.strip()
    owner = owner.strip()
    for key, meta in SECTION_MAP.items():
        if title.startswith(key) or key in title:
            return (*meta, owner)
    return None


def _to_progress_pct(raw: Any) -> float | None:
    if raw is None or raw == "":
        return None
    try:
        val = float(raw)
    except (TypeError, ValueError):
        return None
    if 0 <= val <= 1:
        return round(val * 100, 1)
    if 1 < val <= 100:
        return round(val, 1)
    return None


def _status_from_progress(pct: float | None, *, has_progress: bool) -> str:
    if not has_progress:
        return "attention"
    if pct is None:
        return "attention"
    if pct >= 100:
        return "completed"
    if pct < 30:
        return "attention"
    return "ongoing"


def _is_header_row(a: Any, b: Any) -> bool:
    if isinstance(a, str) and a.strip().startswith("编号"):
        return True
    if isinstance(b, str) and b.strip() in {"内容", "内容（不局限于这些）"}:
        return True
    return False


def _is_number_cell(v: Any) -> bool:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return True
    if isinstance(v, str) and v.strip().isdigit():
        return True
    return False


def parse_key_plan_xlsx(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    groups_order: list[str] = []
    groups: dict[str, dict[str, Any]] = {}
    current: tuple[str, str, str, str] | None = None
    current_task: dict[str, Any] | None = None

    def ensure_group(meta: tuple[str, str, str, str]) -> dict[str, Any]:
        cat, title, subtitle, owner = meta
        if cat not in groups:
            groups[cat] = {
                "id": cat,
                "title": title,
                "subtitle": subtitle,
                "owner": owner,
                "defaultExpanded": False,
                "metrics": [],
            }
            groups_order.append(cat)
        elif owner and not groups[cat].get("owner"):
            groups[cat]["owner"] = owner
        return groups[cat]

    def start_task(
        name: str,
        detail: Any,
        quant: Any,
        target: Any,
        progress: Any,
        deadline: Any,
        notes: Any,
    ) -> None:
        nonlocal current_task
        assert current is not None
        cat, title, _subtitle, owner = current
        g = ensure_group(current)
        pct = _to_progress_pct(progress)
        task = {
            "id": f"{cat}-{len(g['metrics']) + 1}",
            "name": name.strip(),
            "category": cat,
            "taskType": title,
            "projectLevel": "学院重点",
            "majorDirection": title,
            "target": str(target).strip() if target not in (None, "") else "",
            "actual": "",
            "unit": "",
            "progress": 0,
            "status": "attention",
            "owner": owner or title,
            "deadline": str(deadline).strip() if deadline not in (None, "") else "",
            "milestone": str(notes).strip() if notes not in (None, "") else "",
            "materials": [],
            "subitems": [str(detail).strip()] if detail not in (None, "") else [],
            "_progresses": [pct] if pct is not None else [],
            "_quant": str(quant).strip() if quant not in (None, "") else "",
        }
        g["metrics"].append(task)
        current_task = task

    def absorb_row(detail: Any, target: Any, progress: Any, deadline: Any, notes: Any) -> None:
        if current_task is None:
            return
        if detail not in (None, ""):
            text = str(detail).strip()
            if text and text not in current_task["subitems"]:
                current_task["subitems"].append(text)
        pct = _to_progress_pct(progress)
        if pct is not None:
            current_task["_progresses"].append(pct)
        if target not in (None, "") and not current_task["target"]:
            current_task["target"] = str(target).strip()
        if deadline not in (None, "") and not current_task["deadline"]:
            current_task["deadline"] = str(deadline).strip()
        if notes not in (None, ""):
            note = str(notes).strip()
            if note:
                if current_task["milestone"]:
                    current_task["milestone"] = f"{current_task['milestone']}；{note}"
                else:
                    current_task["milestone"] = note

    for r in range(1, ws.max_row + 1):
        a_raw = _raw(ws, r, 1)
        b_raw = _raw(ws, r, 2)
        a = _merged(ws, r, 1)
        b = _merged(ws, r, 2)
        c = _merged(ws, r, 3)
        d = _merged(ws, r, 4)
        e = _raw(ws, r, 5)
        f = _raw(ws, r, 6)
        g = _raw(ws, r, 7)
        h = _raw(ws, r, 8)

        if isinstance(a_raw, str) and a_raw.strip():
            meta = _parse_section(a_raw)
            if meta:
                current = meta
                current_task = None
                continue

        if _is_header_row(a_raw, b_raw):
            continue

        if not current:
            continue

        # 仅当本行 A 格真实写了编号时，开启新任务（避免合并编号重复建任务）
        if _is_number_cell(a_raw):
            name = str(b).strip() if b not in (None, "") else ""
            if not name:
                current_task = None
                continue
            start_task(name, c, d, e, f, g, h)
            continue

        # 人才培养等：无编号、但本行 B 真实写了任务名
        if (
            b_raw not in (None, "")
            and isinstance(b_raw, str)
            and b_raw.strip()
            and b_raw.strip() not in {"内容", "内容（不局限于这些）"}
        ):
            start_task(str(b_raw).strip(), c, d, e, f, g, h)
            continue

        # 无编号行：吞入当前任务（细项 / 进展）
        if any(v not in (None, "") for v in (c, d, e, f, g, h)):
            absorb_row(c, e, f, g, h)

    metrics: list[dict[str, Any]] = []
    for cat in groups_order:
        group = groups[cat]
        for task in group["metrics"]:
            progresses: list[float] = task.pop("_progresses", [])
            task.pop("_quant", None)
            has_progress = bool(progresses)
            if has_progress:
                avg = int(round(sum(progresses) / len(progresses)))
                task["progress"] = avg
                if not task["actual"]:
                    task["actual"] = f"{avg}%"
            else:
                task["progress"] = 0
            task["status"] = _status_from_progress(
                task["progress"] if has_progress else None,
                has_progress=has_progress,
            )
            if task["subitems"] and not task["milestone"]:
                task["milestone"] = "；".join(task["subitems"][:3])
            task["materials"] = list(task["subitems"][:8])
            task.pop("subitems", None)
            metrics.append(task)

    total = len(metrics)
    completed = sum(1 for m in metrics if m["status"] == "completed")
    attention = sum(1 for m in metrics if m["status"] == "attention")
    ongoing = max(total - completed - attention, 0)
    completion_rate = round(sum(m["progress"] for m in metrics) / total) if total else 0

    return {
        "year": "2025",
        "overview": {
            "total": total,
            "completed": completed,
            "ongoing": ongoing,
            "attention": attention,
            "completionRate": completion_rate,
        },
        "groups": [groups[c] for c in groups_order],
        "metrics": metrics,
        "sourceFile": path.name,
    }


if __name__ == "__main__":
    import json
    import sys

    src = Path(sys.argv[1] if len(sys.argv) > 1 else r"d:\UGit\data\学院重点工作进展.xlsx")
    data = parse_key_plan_xlsx(src)
    print(json.dumps(data["overview"], ensure_ascii=False, indent=2))
    for g in data["groups"]:
        print(f"\n## {g['title']} ({g['id']}) owner={g.get('owner')} n={len(g['metrics'])}")
        for m in g["metrics"]:
            print(
                f"  - {m['name']}: {m['progress']}% [{m['status']}] "
                f"target={m['target'] or '-'} milestone={ (m['milestone'] or '')[:40] }"
            )
