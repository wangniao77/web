"""从教职工通讯录 Excel 同步可用字段到 teachers。

可用数据（入库）:
  - 姓名 / 工号 / 性别
  - 系所（部门分组）
  - 职称 / 岗位（从「职务、职称」+ 备注拆分）
  - 在职状态（在职 / 调离 / 退休 / 离世）

不入库（隐私 / 无模型字段）:
  - 手机、办公室电话、邮箱、政治面貌

用法:
  cd backend
  python scripts/import_teacher_contact.py
  python scripts/import_teacher_contact.py --path "D:/UGit/data/8_5/xxx.xlsx" --dry-run
"""

from __future__ import annotations

import argparse
import asyncio
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
from tortoise import Tortoise

BACKEND = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND))

from core.database import TORTOISE_ORM
from Utils.DB.Models.college_ext_models import ResearchPlatform, Teacher
from Utils.DB.read.college_db import resolve_college

SOURCE_TAG = "contact_2026_06"
DEFAULT_PATH = Path(
    r"d:\UGit\data\8_5\大数据与人工智能学院教职工通讯录（2026.6.29更新）.xlsx"
)

# 学术职称（优先写入 title）
ACADEMIC_TITLES = (
    "教授级高级工程师",
    "高级实验师",
    "高级工程师",
    "助理研究员",
    "助理工程师",
    "助理教授",
    "副研究员",
    "副教授",
    "教授",
    "研究员",
    "实验师",
    "工程师",
    "讲师",
    "助教",
    "未定职级",
    "未评级",
)

# 行政/教辅岗位关键词（写入 position）
ADMIN_HINTS = (
    "党委书记",
    "党委副书记",
    "副书记",
    "院长助理",
    "副院长",
    "院长",
    "办公室主任",
    "系主任",
    "副主任",
    "主任",
    "辅导员",
    "教学秘书",
    "学科秘书",
    "科研秘书",
    "秘书",
    "实验员",
    "组织员",
    "科员",
)

TITLE_LEVEL_MAP = {
    "教授": "正高级",
    "教授级高级工程师": "正高级",
    "研究员": "正高级",
    "副教授": "副高级",
    "副研究员": "副高级",
    "高级工程师": "副高级",
    "高级实验师": "副高级",
    "讲师": "中级",
    "工程师": "中级",
    "实验师": "中级",
    "助理研究员": "中级",
    "助教": "初级",
    "助理工程师": "初级",
    "助理教授": "中级",
    "未定职级": None,
    "未评级": None,
}


def _norm_no(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip().replace(".0", "")
    if s.upper() in {"#VALUE!", "NAN", "NONE", "NONE"}:
        return ""
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _norm_title(title: str | None) -> str | None:
    if not title:
        return None
    t = title.strip()
    if t == "未定职级":
        return "未评级"
    return t


def _split_title_position(raw: str | None, note: str | None) -> tuple[str | None, str | None]:
    """从「职务、职称」与备注拆出学术职称 / 行政岗位。"""
    raw = (raw or "").strip()
    note = (note or "").strip()
    parts = [p.strip() for p in re.split(r"[、，,/]", raw) if p and p.strip()]

    academic: list[str] = []
    admin: list[str] = []
    for p in parts:
        matched_acad = None
        for a in ACADEMIC_TITLES:
            if a in p:
                matched_acad = a
                break
        if matched_acad and p == matched_acad:
            academic.append(matched_acad)
        elif matched_acad and any(h in p for h in ADMIN_HINTS):
            # 如「副院长、副教授」已被拆分；若单段混写则两边都收
            academic.append(matched_acad)
            admin.append(p)
        elif matched_acad:
            academic.append(matched_acad)
        else:
            admin.append(p)

    # 备注中的系主任 / 课程负责人等补入岗位
    if note:
        for hint in (
            "系主任",
            "副主任",
            "主任",
            "支部书记",
            "组织委员",
            "宣传委员",
            "课程负责人",
            "教学团队负责人",
            "团队负责人",
            "专业负责人",
            "院长助理",
            "组织员",
            "工会",
        ):
            if hint in note and note not in admin:
                admin.append(note)
                break
        else:
            # 无明确 hint 时若备注较短也并入岗位补充
            if note and len(note) <= 40 and note not in admin:
                # 仅当 admin 为空时用备注，避免噪音
                if not admin:
                    admin.append(note)

    title = _norm_title(academic[0] if academic else None)
    # 纯行政岗且无学术职称时：辅导员/秘书等可作为 title 也可作 position
    position = "，".join(dict.fromkeys(admin)) if admin else None
    if title is None and position:
        # 办公室类岗位：title 用岗位名（与现有花名册「辅导员」习惯一致）
        for prefer in ("辅导员", "教学秘书", "学科秘书", "办公室主任", "实验员", "秘书"):
            if prefer in position:
                title = prefer
                break
        if title is None and len(admin) == 1:
            title = admin[0]
    return title, position


def _dept_from_section(section: str | None, note: str | None) -> str | None:
    if not section:
        return None
    s = section.strip()
    if s == "党政班子成员":
        note = note or ""
        if "大数据管理与应用" in note:
            return "大数据管理与应用系"
        # 仅当备注明确点名系所时覆盖；避免「人工智能通识教学团队」误匹配
        for dept_name in (
            "计算机科学与技术系",
            "软件工程系",
            "人工智能系",
            "电子商务系",
            "大数据管理与应用系",
        ):
            short = dept_name.replace("系", "")
            if dept_name in note or f"{short}系" in note:
                return dept_name
        return "学院党政"
    if s.startswith("办公室"):
        return "办公室"
    return s.split("（")[0].strip()


def _parse_active(ws) -> list[dict[str, Any]]:
    dept = None
    people: list[dict[str, Any]] = []
    for r in range(4, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if isinstance(b, str) and b.strip():
            first = b.split("\n")[0].strip()
            if "调离" in first or "退休" in first or "概况" in first:
                break
            if any(k in first for k in ("系", "班子", "办公室", "中心", "辅导", "实验", "行政")):
                dept = first
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        try:
            seq = int(c) if c is not None else None
        except (TypeError, ValueError):
            seq = None
        if seq is None or not isinstance(d, str) or not (1 < len(d.strip()) <= 10):
            continue
        name = d.strip()
        no = _norm_no(ws.cell(r, 5).value)
        gender = str(ws.cell(r, 6).value or "").strip() or None
        title_raw = str(ws.cell(r, 7).value or "").strip() or None
        note = str(ws.cell(r, 12).value or "").strip() or None
        title, position = _split_title_position(title_raw, note)
        department = _dept_from_section(dept, note)
        people.append(
            {
                "name": name,
                "teacher_no": no or None,
                "gender": gender,
                "title": title,
                "title_level": TITLE_LEVEL_MAP.get(title or "", None),
                "department": department,
                "position": position,
                "status": "active",
                "note": note,
                "title_raw": title_raw,
            }
        )
    return people


def _parse_status_section(ws) -> list[dict[str, Any]]:
    """解析调离 / 退休 / 离世名单。列：B姓名 C工号 E性别 F职务职称 K备注。"""
    mode: str | None = None
    out: list[dict[str, Any]] = []
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if not isinstance(b, str):
            continue
        first = b.strip()
        if first == "调离教师":
            mode = "transferred"
            continue
        if first == "退休教师":
            mode = "retired"
            continue
        if mode is None:
            continue
        if first in {"姓名", "调离教师", "退休教师"} or "汇总表" in first:
            continue
        # 人名行：下一列是工号
        no = _norm_no(ws.cell(r, 3).value)
        if not no:
            continue
        name = first
        gender = str(ws.cell(r, 5).value or "").strip() or None
        title_raw = str(ws.cell(r, 6).value or "").strip() or None
        note = str(ws.cell(r, 11).value or "").strip() or None
        title, position = _split_title_position(title_raw, None)
        status = mode
        if note and ("离世" in note or "去世" in note):
            status = "deceased"
        out.append(
            {
                "name": name,
                "teacher_no": no,
                "gender": gender,
                "title": title,
                "title_level": TITLE_LEVEL_MAP.get(title or "", None),
                "department": None,
                "position": position,
                "status": status,
                "note": note,
                "title_raw": title_raw,
            }
        )
    return out


def _parse_sheet1_nos(wb) -> dict[str, str]:
    if "Sheet1" not in wb.sheetnames:
        return {}
    ws = wb["Sheet1"]
    mapping: dict[str, str] = {}
    for r in range(1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        no = _norm_no(ws.cell(r, 2).value)
        if name and no:
            mapping[str(name).strip()] = no
    return mapping


def _parse_teams(wb) -> list[dict[str, Any]]:
    """从「学院领军人才、团队等」提取团队 → research_platforms。"""
    sheet = "学院领军人才、团队等"
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    teams: list[dict[str, Any]] = []
    section: str | None = None
    leader: str | None = None
    members: list[str] = []

    def _flush() -> None:
        nonlocal section, leader, members
        if not section or "团队" not in section:
            section, leader, members = None, None, []
            return
        name = re.sub(r"\s*\d+人\s*$", "", section).strip()
        m = re.search(r"(\d+)\s*人", section)
        count = int(m.group(1)) if m else (len(members) or None)
        if not leader and members:
            leader = members[0]
        if name and leader:
            teams.append(
                {
                    "name": name,
                    "category": "团队",
                    "leader": leader,
                    "member_count": count,
                }
            )
        section, leader, members = None, None, []

    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        note = ws.cell(r, 10).value
        if isinstance(b, str) and c is None and b.strip():
            _flush()
            section = b.strip()
            leader = None
            members = []
            continue
        if not (isinstance(b, str) and c is not None):
            continue
        name = b.strip()
        members.append(name)
        note_s = str(note or "").strip()
        if leader is None and (
            note_s == "负责人"
            or note_s.startswith("负责人")
            or "团队统筹负责人" in note_s
        ):
            leader = name
    _flush()
    return teams


def _should_overwrite_title(old: str | None, new: str | None) -> bool:
    if not new:
        return False
    if not old:
        return True
    weak = {"未评级", "未定职级", "辅导员", "教学秘书", "学科秘书", "秘书", "实验员", "办公室主任"}
    if new in weak and old not in weak:
        # 不用「未定职级」覆盖已有明确职称
        return False
    if old in weak and new not in weak:
        return True
    if old == new:
        return False
    acad_set = set(ACADEMIC_TITLES) | {"未评级"}
    if old in acad_set and new in acad_set:
        return True
    if old not in acad_set and new in acad_set:
        return True
    return False


def _merge_position(old: str | None, new: str | None) -> str | None:
    if not new:
        return old
    if not old or old in {"教师", "科研人员"}:
        return new
    if new in old:
        return old
    if old in new:
        return new
    # 双方都有实质内容：保留信息量更大的一侧
    if len(old) >= len(new) + 4:
        return old
    return new


async def _find_teacher(college_id: int, *, name: str, teacher_no: str | None) -> Teacher | None:
    if teacher_no:
        t = await Teacher.get_or_none(college_id=college_id, teacher_no=teacher_no)
        if t:
            return t
    return await Teacher.get_or_none(college_id=college_id, name=name)


async def _upsert_teacher(
    college,
    row: dict[str, Any],
    *,
    dry_run: bool,
    fill_no_only: bool = False,
) -> str:
    existing = await _find_teacher(college.id, name=row["name"], teacher_no=row.get("teacher_no"))
    if fill_no_only:
        if existing and not existing.teacher_no and row.get("teacher_no"):
            if not dry_run:
                existing.teacher_no = row["teacher_no"]
                src = existing.source or ""
                if SOURCE_TAG not in src.split(","):
                    existing.source = f"{src},{SOURCE_TAG}" if src else SOURCE_TAG
                await existing.save()
            return "no_filled"
        return "skip"

    payload: dict[str, Any] = {
        "name": row["name"],
        "status": row["status"],
    }
    if row.get("teacher_no"):
        payload["teacher_no"] = row["teacher_no"]
    if row.get("gender"):
        payload["gender"] = row["gender"]
    if row.get("department"):
        payload["department"] = row["department"]

    if existing:
        changed = False
        # title
        new_title = row.get("title")
        if _should_overwrite_title(existing.title, new_title):
            existing.title = new_title
            lvl = row.get("title_level")
            if lvl:
                existing.title_level = lvl
            changed = True
        # position
        merged_pos = _merge_position(existing.position, row.get("position"))
        if merged_pos and merged_pos != existing.position:
            existing.position = merged_pos
            changed = True
        for k, v in payload.items():
            if v is None or v == "":
                continue
            if getattr(existing, k) != v:
                setattr(existing, k, v)
                changed = True
        # 在职名单优先：若此人在 active 列表，强制 active
        if row["status"] == "active" and existing.status != "active":
            existing.status = "active"
            changed = True
        src = existing.source or ""
        if SOURCE_TAG not in src.split(","):
            existing.source = f"{src},{SOURCE_TAG}" if src else SOURCE_TAG
            changed = True
        if changed:
            if not dry_run:
                await existing.save()
            return "updated"
        return "skip"

    # 调离/退休人员若不在库中：仍建档便于留痕
    create_payload = {
        "name": row["name"],
        "teacher_no": row.get("teacher_no"),
        "gender": row.get("gender"),
        "title": row.get("title"),
        "title_level": row.get("title_level"),
        "department": row.get("department"),
        "position": row.get("position"),
        "status": row["status"],
        "source": SOURCE_TAG,
    }
    if not dry_run:
        await Teacher.create(college=college, **create_payload)
    return "created"


async def _upsert_team(college, team: dict[str, Any], *, dry_run: bool) -> str:
    existing = await ResearchPlatform.get_or_none(
        college_id=college.id, name=team["name"], category="团队"
    )
    if existing:
        changed = False
        for k in ("leader", "member_count"):
            v = team.get(k)
            if v is not None and getattr(existing, k) != v:
                setattr(existing, k, v)
                changed = True
        if team.get("source_file") and existing.source_file != team["source_file"]:
            existing.source_file = team["source_file"]
            changed = True
        if changed and not dry_run:
            await existing.save()
        return "updated" if changed else "skip"
    if not dry_run:
        await ResearchPlatform.create(
            college=college,
            name=team["name"],
            category=team["category"],
            leader=team.get("leader"),
            member_count=team.get("member_count"),
            source_file=team.get("source_file"),
        )
    return "created"


async def main() -> None:
    parser = argparse.ArgumentParser(description="导入教职工通讯录可用字段")
    parser.add_argument("--path", type=Path, default=DEFAULT_PATH)
    parser.add_argument("--college-code", default="big-data-ai")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--skip-teams", action="store_true", help="不导入团队平台")
    args = parser.parse_args()

    if not args.path.exists():
        raise SystemExit(f"file not found: {args.path}")

    wb = openpyxl.load_workbook(args.path, data_only=True)
    if "全院通讯录" not in wb.sheetnames:
        raise SystemExit(f"missing sheet 全院通讯录, got {wb.sheetnames}")

    ws = wb["全院通讯录"]
    active = _parse_active(ws)
    leavers = _parse_status_section(ws)
    sheet1_nos = _parse_sheet1_nos(wb)
    teams = [] if args.skip_teams else _parse_teams(wb)
    for t in teams:
        t["source_file"] = args.path.name

    active_names = {p["name"] for p in active}
    active_nos = {p["teacher_no"] for p in active if p.get("teacher_no")}

    # 调离/退休中若仍出现在在职名单，以在职为准
    filtered_leavers = []
    for row in leavers:
        if row["name"] in active_names or (
            row.get("teacher_no") and row["teacher_no"] in active_nos
        ):
            continue
        filtered_leavers.append(row)

    print(
        f"parsed active={len(active)} leavers={len(leavers)} "
        f"(apply={len(filtered_leavers)}) sheet1={len(sheet1_nos)} teams={len(teams)}"
    )
    if args.dry_run:
        print("[dry-run] no DB writes")

    await Tortoise.init(config=TORTOISE_ORM, _enable_global_fallback=True)
    college = await resolve_college(args.college_code)
    if not college:
        raise SystemExit(f"college not found: {args.college_code}")
    print(f"college={college.code} id={college.id}")

    stats = {
        "active_created": 0,
        "active_updated": 0,
        "active_skip": 0,
        "leaver_created": 0,
        "leaver_updated": 0,
        "leaver_skip": 0,
        "no_filled": 0,
        "team_created": 0,
        "team_updated": 0,
        "team_skip": 0,
    }

    for row in active:
        action = await _upsert_teacher(college, row, dry_run=args.dry_run)
        if action == "created":
            stats["active_created"] += 1
        elif action == "updated":
            stats["active_updated"] += 1
        else:
            stats["active_skip"] += 1

    for row in filtered_leavers:
        action = await _upsert_teacher(college, row, dry_run=args.dry_run)
        if action == "created":
            stats["leaver_created"] += 1
        elif action == "updated":
            stats["leaver_updated"] += 1
        else:
            stats["leaver_skip"] += 1

    # Sheet1：仅为已有教师补全工号
    for name, no in sheet1_nos.items():
        action = await _upsert_teacher(
            college,
            {"name": name, "teacher_no": no, "status": "active"},
            dry_run=args.dry_run,
            fill_no_only=True,
        )
        if action == "no_filled":
            stats["no_filled"] += 1

    if not args.skip_teams:
        for team in teams:
            action = await _upsert_team(college, team, dry_run=args.dry_run)
            stats[f"team_{action}"] += 1

    print("RESULTS", stats)

    # 汇总核对
    from collections import Counter

    teachers = await Teacher.filter(college_id=college.id)
    print("status", dict(Counter(t.status for t in teachers)))
    print("department", dict(Counter(t.department for t in teachers if t.status == "active")))
    print("IMPORT_TEACHER_CONTACT_OK")
    await Tortoise.close_connections()


if __name__ == "__main__":
    asyncio.run(main())
