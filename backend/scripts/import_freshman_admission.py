"""导入 2026 新生录取基础表（筛选可用列）→ students / student_admission / enrollment_cohorts。

数据特点：
  - 尚无正式学号 → 临时业务键 KSH{考生号}（考生号全表唯一）
  - status=admitted（已录取未入正式学籍），避免污染在校 active 花名册口径
  - 补最小 academic_snapshots(grade=入学年)，使生源分析视图可关联

入库字段（有用）：
  年份、省份、校区、姓名、性别、校标专业、学历、投档成绩/总分、
  考生号、招生类别(港澳台)、录取专业志愿(聚合一志愿率)、政治面貌

明确不入库（隐私/无模型/无分析价值）：
  身份证、电话/手机、家庭地址、收件人、邮编、出生年月、
  身高体重病史、志愿 1–20 明细、各类国标/校标代码列

用法:
  cd backend
  python scripts/import_freshman_admission.py
  python scripts/import_freshman_admission.py --dry-run
"""

from __future__ import annotations

import argparse
import asyncio
import re
import sys
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from tortoise import Tortoise

BACKEND = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND))

from core.database import TORTOISE_ORM
from Utils.DB.Models.college_ext_models import EnrollmentCohort
from Utils.DB.Models.college_student_models import (
    AcademicSnapshot,
    StudentAdmission,
    StudentProfile,
)
from Utils.DB.Models.major_models import Major
from Utils.DB.read.college_db import resolve_college

SOURCE_TAG = "freshman_admission_2026"
DEFAULT_PATH = Path(
    r"d:\UGit\data\8_5\学生基础表-全字段模版-2026年-大数据与人工智能学院-四年"
    r"\学生基础表-全字段模版-2026年-大数据与人工智能学院-四年.xlsx"
)

# Excel 校标专业名 → 库内 majors.name
MAJOR_ALIASES: dict[str, str] = {
    "人工智能": "人工智能",
    "软件工程": "软件工程",
    "计算机科学与技术": "计算机科学与技术",
    "计算机科学与技术（创新实验班）": "计算机科学与技术(实验区)",
    "计算机科学与技术（中外联合培养项目）": "计算机科学与技术(中外联合培养项目班)",
}

KEEP_HEADERS = {
    "年份",
    "省份",
    "校区",
    "入学类型",
    "招生类别",
    "校标批次名称",
    "校标科类名称",
    "考生号",
    "姓名",
    "性别名称",
    "学院名称",
    "录取专业志愿",
    "校标专业",
    "打印专业",
    "学历名称",
    "学制名称",
    "政治面貌",
    "民族名称",
    "毕业中学",
    "特征成绩",
    "投档成绩",
    "投档成绩（整数）",
    "总分",
    "位次",
    "选考科目",
    "国标专业名称",
    "计划性质（国标）",
    "考生类别名称",
    "外语语种名称",
    "考试类型名称",
    "报到校区",
    "通知书编号",
}


def _cell(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _to_int(v: Any) -> int | None:
    s = _cell(v)
    if not s:
        return None
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return None


def _to_decimal(v: Any) -> Decimal | None:
    s = _cell(v)
    if not s:
        return None
    try:
        d = Decimal(s)
    except (InvalidOperation, ValueError):
        return None
    if d == 0:
        # 港澳台联招等可能把「考生成绩」填 0，改用投档/总分；此处 0 视为无效
        return None
    return d


def _provisional_student_no(candidate_no: str) -> str:
    return f"KSH{candidate_no}"


def _hmt_status(*, province: str, enroll_category: str, enroll_type: str) -> str | None:
    blob = f"{province}|{enroll_category}|{enroll_type}"
    if any(k in blob for k in ("港澳台", "香港", "澳门", "台湾")):
        if "香港" in province:
            return "香港"
        if "澳门" in province:
            return "澳门"
        if "台湾" in province:
            return "台湾"
        return "港澳台"
    return None


def _is_first_choice(vol: str) -> bool:
    v = (vol or "").strip()
    return v in {"一志愿", "1", "第一志愿"}


def _pick_score(row: dict[str, Any]) -> Decimal | None:
    for key in ("投档成绩", "投档成绩（整数）", "总分", "特征成绩"):
        d = _to_decimal(row.get(key))
        if d is not None:
            return d
    return None


def _read_rows(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    headers: list[str] | None = None
    out: list[dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            headers = [_cell(h) or f"c{j}" for j, h in enumerate(row)]
            continue
        assert headers is not None
        raw = dict(zip(headers, row))
        name = _cell(raw.get("姓名"))
        cand = _cell(raw.get("考生号"))
        if not name or not cand:
            continue
        # 仅保留有用列，降低后续处理噪音
        slim = {k: raw.get(k) for k in KEEP_HEADERS if k in raw}
        slim["姓名"] = name
        slim["考生号"] = cand
        out.append(slim)
    wb.close()
    return out


async def _ensure_major(college_id: int, excel_name: str) -> Major | None:
    name = (excel_name or "").strip()
    if not name:
        return None
    mapped = MAJOR_ALIASES.get(name, name)
    major = await Major.get_or_none(college_id=college_id, name=mapped)
    if major:
        return major
    # 宽松：去括号差异再找
    for m in await Major.filter(college_id=college_id):
        a = re.sub(r"[\s（）()]", "", m.name)
        b = re.sub(r"[\s（）()]", "", mapped)
        if a == b or a in b or b in a:
            return m
    code = re.sub(r"\W+", "", mapped)[:32] or f"m{college_id}"
    return await Major.create(college_id=college_id, name=mapped, code=code)


async def _upsert_student(
    college,
    row: dict[str, Any],
    *,
    dry_run: bool,
) -> str:
    cand = _cell(row.get("考生号"))
    student_no = _provisional_student_no(cand)
    name = _cell(row.get("姓名"))
    year = _to_int(row.get("年份")) or 2026
    major_excel = _cell(row.get("校标专业")) or _cell(row.get("打印专业"))
    major = await _ensure_major(college.id, major_excel)
    province = _cell(row.get("省份"))
    score = _pick_score(row)
    campus = _cell(row.get("报到校区")) or _cell(row.get("校区")) or None
    gender = _cell(row.get("性别名称")) or None
    edu = _cell(row.get("学历名称")) or None
    political = _cell(row.get("政治面貌")) or None
    hmt = _hmt_status(
        province=province,
        enroll_category=_cell(row.get("招生类别")),
        enroll_type=_cell(row.get("入学类型")),
    )

    existing = await StudentProfile.get_or_none(student_no=student_no)
    profile_payload = {
        "name": name,
        "gender": gender,
        "status": "admitted",
        "college_id": college.id,
        "major_id": major.id if major else None,
        "campus": campus,
        "education_level": edu,
        "enrollment_year": year,
        "teaching_department": _cell(row.get("学院名称")) or college.name,
        "major_name": (major.name if major else major_excel) or None,
        "political_status": political,
    }

    if existing:
        changed = False
        for k, v in profile_payload.items():
            if v is None or v == "":
                continue
            if getattr(existing, k) != v:
                setattr(existing, k, v)
                changed = True
        if not dry_run and changed:
            await existing.save()
        profile = existing
        action = "updated" if changed else "skip"
    else:
        if not dry_run:
            profile = await StudentProfile.create(student_no=student_no, **profile_payload)
        else:
            profile = None  # type: ignore[assignment]
        action = "created"

    if dry_run:
        return action

    assert profile is not None
    adm = await StudentAdmission.get_or_none(student_id=profile.id)
    adm_payload = {
        "admission_score": score,
        "source_place": province or None,
        "hmt_status": hmt,
    }
    if adm:
        for k, v in adm_payload.items():
            if v is None:
                continue
            setattr(adm, k, v)
        await adm.save()
    else:
        await StudentAdmission.create(student=profile, **adm_payload)

    snap = await AcademicSnapshot.get_or_none(student_id=profile.id, grade=year)
    snap_payload = {
        "college_id": college.id,
        "major_id": major.id if major else None,
        "major_name": profile.major_name,
        "academic_year": f"{year}-{year + 1}",
    }
    if snap:
        for k, v in snap_payload.items():
            if v is not None:
                setattr(snap, k, v)
        await snap.save()
    else:
        await AcademicSnapshot.create(student=profile, grade=year, **snap_payload)

    return action


async def _upsert_cohorts(
    college,
    rows: list[dict[str, Any]],
    *,
    dry_run: bool,
) -> dict[str, int]:
    buckets: dict[tuple[int, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        year = _to_int(row.get("年份")) or 2026
        major_excel = _cell(row.get("校标专业")) or _cell(row.get("打印专业")) or "未知"
        buckets[(year, major_excel)].append(row)

    stats = {"created": 0, "updated": 0, "skip": 0}
    for (year, major_excel), group in buckets.items():
        major = await _ensure_major(college.id, major_excel)
        if not major:
            stats["skip"] += 1
            continue
        scores: list[Decimal] = []
        ranks: list[int] = []
        first_n = 0
        for r in group:
            sc = _pick_score(r)
            if sc is not None:
                scores.append(sc)
            rk = _to_int(r.get("位次"))
            if rk and rk > 0:
                ranks.append(rk)
            if _is_first_choice(_cell(r.get("录取专业志愿"))):
                first_n += 1
        n = len(group)
        avg = (sum(scores) / Decimal(len(scores))).quantize(Decimal("0.01")) if scores else None
        mn = min(scores) if scores else None
        rate = (Decimal(first_n) / Decimal(n)).quantize(Decimal("0.0001")) if n else None
        cutoff = max(ranks) if ranks else None  # 位次越大越靠后；录取线常用最差位次

        existing = await EnrollmentCohort.get_or_none(
            college_id=college.id, major_id=major.id, year=year
        )
        payload = {
            "enrolled_count": n,
            "first_choice_rate": rate,
            "avg_score": avg,
            "min_score": mn,
            "rank_cutoff": cutoff,
            "source": SOURCE_TAG,
        }
        if existing:
            changed = False
            for k, v in payload.items():
                if getattr(existing, k) != v:
                    setattr(existing, k, v)
                    changed = True
            if changed and not dry_run:
                await existing.save()
            stats["updated" if changed else "skip"] += 1
        else:
            if not dry_run:
                await EnrollmentCohort.create(
                    college=college,
                    major=major,
                    year=year,
                    **payload,
                )
            stats["created"] += 1
    return stats


async def main() -> None:
    parser = argparse.ArgumentParser(description="导入新生录取基础表（筛选列）")
    parser.add_argument("--path", type=Path, default=DEFAULT_PATH)
    parser.add_argument("--college-code", default="big-data-ai")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.path.exists():
        raise SystemExit(f"file not found: {args.path}")

    rows = _read_rows(args.path)
    print(f"parsed rows={len(rows)} useful_headers={len(KEEP_HEADERS)}")
    if args.dry_run:
        print("[dry-run] no DB writes")

    await Tortoise.init(config=TORTOISE_ORM, _enable_global_fallback=True)
    college = await resolve_college(args.college_code)
    if not college:
        raise SystemExit(f"college not found: {args.college_code}")
    print(f"college={college.code} id={college.id}")

    stats = {"created": 0, "updated": 0, "skip": 0}
    for row in rows:
        action = await _upsert_student(college, row, dry_run=args.dry_run)
        stats[action] = stats.get(action, 0) + 1

    cohort_stats = await _upsert_cohorts(college, rows, dry_run=args.dry_run)
    print("STUDENTS", stats)
    print("COHORTS", cohort_stats)

    admitted = await StudentProfile.filter(
        college_id=college.id, status="admitted", enrollment_year=2026
    ).count()
    cohorts = await EnrollmentCohort.filter(college_id=college.id, year=2026).count()
    print(f"verify admitted_2026={admitted} cohorts_2026={cohorts}")
    print("IMPORT_FRESHMAN_ADMISSION_OK")
    await Tortoise.close_connections()


if __name__ == "__main__":
    asyncio.run(main())
