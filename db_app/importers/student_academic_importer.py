from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from app.models import StudentAcademicRecord


COLUMN_TO_FIELD = {
    "年级": "grade",
    "学号": "student_id",
    "姓名": "name",
    "姓名拼音": "name_pinyin",
    "性别": "gender",
    "上课院系": "teaching_department",
    "专业名称": "major_name",
    "专业方向名称": "major_direction_name",
    "班级": "class_name",
    "校区": "campus",
    "招生学科": "admission_subject",
    "辅修专业名称": "minor_major_name",
    "出生日期": "birth_date",
    "入学年份": "enrollment_year",
    "培养层次": "education_level",
    "研究方向": "research_direction",
    "导师姓名": "supervisor_name",
    "民族": "ethnicity",
    "籍贯": "native_place",
    "政治面貌": "political_status",
    "港澳台侨外": "hongkong_macao_taiwan_overseas",
    "健康状况": "health_status",
    "电话": "phone",
    "体育达标": "pe_standard",
    "选课组": "course_selection_group",
    "专业课_课程数": "major_course_count",
    "专业课_平均分": "major_course_avg_score",
    "专业课_最高分": "major_course_max_score",
    "专业课_最低分": "major_course_min_score",
    "专业课_方差": "major_course_variance",
    "学科基础课_课程数": "subject_basic_course_count",
    "学科基础课_平均分": "subject_basic_course_avg_score",
    "学科基础课_最高分": "subject_basic_course_max_score",
    "学科基础课_最低分": "subject_basic_course_min_score",
    "学科基础课_方差": "subject_basic_course_variance",
    "通识课_课程数": "general_course_count",
    "通识课_平均分": "general_course_avg_score",
    "通识课_最高分": "general_course_max_score",
    "通识课_最低分": "general_course_min_score",
    "通识课_方差": "general_course_variance",
    "必修_课程数": "required_course_count",
    "必修_平均分": "required_course_avg_score",
    "必修_最高分": "required_course_max_score",
    "必修_最低分": "required_course_min_score",
    "必修_方差": "required_course_variance",
    "选修_课程数": "elective_course_count",
    "选修_平均分": "elective_course_avg_score",
    "选修_最高分": "elective_course_max_score",
    "选修_最低分": "elective_course_min_score",
    "选修_方差": "elective_course_variance",
    "全部课程_课程数": "all_course_count",
    "全部课程_平均分": "all_course_avg_score",
    "全部课程_最高分": "all_course_max_score",
    "全部课程_最低分": "all_course_min_score",
    "全部课程_方差": "all_course_variance",
    "缺考次数": "absent_exam_count",
    "补考次数": "makeup_exam_count",
    "重修次数": "retake_count",
    "必修": "required_credits",
    "选修": "elective_credits",
    "其它": "other_credits",
    "重修学分": "retake_credits",
    "获得总学分": "earned_total_credits",
    "不及格总学分": "failed_total_credits",
    "主修总学分": "major_total_credits",
    "辅修总学分": "minor_total_credits",
    "总绩点分": "total_grade_points",
    "平均学分绩点": "average_credit_gpa",
    "考四级次数": "cet4_exam_count",
    "四级": "cet4_score",
    "考六级次数": "cet6_exam_count",
    "六级": "cet6_score",
    "楼栋": "building",
    "宿舍名称": "dormitory_name",
    "班主任": "class_teacher",
    "辅导员": "counselor",
    "学科竞赛获奖次数": "competition_award_count",
    "学科竞赛获奖明细": "competition_award_detail",
}

INT_FIELDS = {
    "grade",
    "enrollment_year",
    "major_course_count",
    "subject_basic_course_count",
    "general_course_count",
    "required_course_count",
    "elective_course_count",
    "all_course_count",
    "absent_exam_count",
    "makeup_exam_count",
    "retake_count",
    "cet4_exam_count",
    "cet6_exam_count",
    "competition_award_count",
}

DECIMAL_FIELDS = {
    "major_course_avg_score",
    "major_course_max_score",
    "major_course_min_score",
    "major_course_variance",
    "subject_basic_course_avg_score",
    "subject_basic_course_max_score",
    "subject_basic_course_min_score",
    "subject_basic_course_variance",
    "general_course_avg_score",
    "general_course_max_score",
    "general_course_min_score",
    "general_course_variance",
    "required_course_avg_score",
    "required_course_max_score",
    "required_course_min_score",
    "required_course_variance",
    "elective_course_avg_score",
    "elective_course_max_score",
    "elective_course_min_score",
    "elective_course_variance",
    "all_course_avg_score",
    "all_course_max_score",
    "all_course_min_score",
    "all_course_variance",
    "required_credits",
    "elective_credits",
    "other_credits",
    "retake_credits",
    "earned_total_credits",
    "failed_total_credits",
    "major_total_credits",
    "minor_total_credits",
    "total_grade_points",
    "average_credit_gpa",
    "cet4_score",
    "cet6_score",
}

DATE_FIELDS = {"birth_date"}


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[str] = field(default_factory=list)


async def import_student_academic_records_from_xlsx(
    xlsx_file: str | Path | BinaryIO,
    *,
    sheet_name: str | None = None,
    batch_size: int = 500,
    update_existing: bool = False,
) -> ImportResult:
    """读取“学籍成绩合并_每人一行.xlsx”，逐行写入 PostgreSQL。

    调用前需要先完成 Tortoise.init()。函数本身不负责创建连接，也不会提交任何真实文件路径。
    """

    workbook = load_workbook(xlsx_file, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    source_file = str(xlsx_file) if isinstance(xlsx_file, (str, Path)) else None

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = _build_header_index(header_row)
    _validate_required_headers(header_index)

    result = ImportResult()
    pending: list[StudentAcademicRecord] = []

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty_row(row):
            result.skipped += 1
            continue

        try:
            payload = _row_to_payload(row, header_index)
            if not payload.get("student_id"):
                result.skipped += 1
                result.errors.append(f"第 {row_number} 行缺少学号，已跳过")
                continue

            payload["source_file"] = source_file
            payload["source_sheet"] = worksheet.title
            payload["row_number"] = row_number

            if update_existing:
                _, created = await StudentAcademicRecord.update_or_create(
                    defaults=_without_unique_keys(payload),
                    grade=payload.get("grade"),
                    student_id=payload["student_id"],
                )
                if created:
                    result.created += 1
                else:
                    result.updated += 1
                continue

            pending.append(StudentAcademicRecord(**payload))
            if len(pending) >= batch_size:
                await StudentAcademicRecord.bulk_create(pending)
                result.created += len(pending)
                pending.clear()
        except Exception as exc:  # noqa: BLE001 - 导入时保留行号比异常类型更重要
            result.errors.append(f"第 {row_number} 行导入失败：{exc}")

    if pending:
        await StudentAcademicRecord.bulk_create(pending)
        result.created += len(pending)

    workbook.close()
    return result


def _build_header_index(header_row: tuple[Any, ...]) -> dict[str, int]:
    return {str(value).strip(): index for index, value in enumerate(header_row) if value is not None}


def _validate_required_headers(header_index: dict[str, int]) -> None:
    missing_headers = [header for header in COLUMN_TO_FIELD if header not in header_index]
    if missing_headers:
        raise ValueError(f"Excel 表头缺失：{', '.join(missing_headers)}")


def _row_to_payload(row: tuple[Any, ...], header_index: dict[str, int]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for column_name, field_name in COLUMN_TO_FIELD.items():
        raw_value = row[header_index[column_name]] if header_index[column_name] < len(row) else None
        payload[field_name] = _normalize_value(field_name, raw_value)
    return payload


def _normalize_value(field_name: str, value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None

    if field_name == "student_id":
        return str(value).strip()
    if field_name in INT_FIELDS:
        return _to_int(value)
    if field_name in DECIMAL_FIELDS:
        return _to_decimal(value)
    if field_name in DATE_FIELDS:
        return _to_date(value)

    return str(value).strip()


def _to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    return int(Decimal(str(value)))


def _to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation as exc:
        raise ValueError(f"无法转换为数字：{value}") from exc


def _to_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    if text.isdigit() and len(text) == 8:
        return datetime.strptime(text, "%Y%m%d").date()
    if text.isdigit():
        return from_excel(int(text)).date()

    return datetime.strptime(text, "%Y-%m-%d").date()


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)


def _without_unique_keys(payload: dict[str, Any]) -> dict[str, Any]:
    return {key: value for key, value in payload.items() if key not in {"grade", "student_id"}}
